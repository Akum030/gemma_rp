import json, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ───────────────────────────────────────────────────────────────────
def load(p):
    with open(p, encoding='utf-8') as f:
        return [json.loads(l) for l in f if l.strip()]

def norm(s): return str(s).lower().strip()

def to_flat(d):
    out = {}
    for k, v in d.items():
        if isinstance(v, list):
            out[norm(k)] = [norm(x) for x in v]
        else:
            out[norm(k)] = norm(v)
    return out

# ── load ──────────────────────────────────────────────────────────────────────
gold = load('bee/gold_1k_v2.jsonl')
v6   = load('bee/v6_flat_results_fixed.jsonl')
qm   = load('bee/qmeans_v2_results.jsonl')

qm_by_q = {r['query'].strip().lower(): to_flat(r.get('attributes', {})) for r in qm}
v6_by_q = {r['query'].strip().lower(): to_flat(r.get('attributes', {})) for r in v6}

# ── styles ────────────────────────────────────────────────────────────────────
GREEN      = PatternFill("solid", fgColor="C6EFCE")
RED        = PatternFill("solid", fgColor="FFC7CE")
YELLOW     = PatternFill("solid", fgColor="FFEB9C")
BLUE_HDR   = PatternFill("solid", fgColor="1F4E79")
QUERY_BG   = PatternFill("solid", fgColor="D6E4F0")
SUMMARY_BG = PatternFill("solid", fgColor="E2EFDA")
GREY       = PatternFill("solid", fgColor="F2F2F2")
WHITE      = PatternFill("solid", fgColor="FFFFFF")

bold_white  = Font(bold=True, color="FFFFFF", size=10)
bold_black  = Font(bold=True, size=10)
bold_green  = Font(bold=True, color="375623", size=10)
bold_red    = Font(bold=True, color="9C0006", size=10)
normal      = Font(size=10)
center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
left        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin        = Side(style='thin', color='BBBBBB')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, fill=None, font=None, align=None):
    if fill:  cell.fill  = fill
    if font:  cell.font  = font
    if align: cell.alignment = align
    cell.border = border

# ── workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Per-Query Comparison"
ws.freeze_panes = "A2"

# column widths
widths = [4, 40, 22, 22, 22, 22, 10, 10, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# header row
headers = ["#", "Query", "Ground Truth (correct answer)",
           "Ground Truth Value", "qmeans Value", "V6 Value",
           "qmeans\nMATCH?", "V6\nMATCH?", "Notes"]
ws.row_dimensions[1].height = 35
for ci, h in enumerate(headers, 1):
    c = ws.cell(1, ci, h)
    style(c, fill=BLUE_HDR, font=bold_white, align=center)

# ── main loop ─────────────────────────────────────────────────────────────────
row = 2
grand_qm_correct = grand_v6_correct = grand_total = 0

for qi, g in enumerate(gold, 1):
    query = g['query']
    gold_attrs = to_flat(g.get('attributes', {}))
    qm_attrs   = qm_by_q.get(query.strip().lower(), {})
    v6_attrs   = v6_by_q.get(query.strip().lower(), {})

    if not gold_attrs:
        continue

    qm_correct = 0
    v6_correct = 0
    total_keys = len(gold_attrs)

    first_row = row

    for key, gold_val in sorted(gold_attrs.items()):
        gold_vals = gold_val if isinstance(gold_val, list) else [gold_val]

        # what did qmeans predict for this key?
        qm_raw = qm_attrs.get(key)
        qm_vals = qm_raw if isinstance(qm_raw, list) else ([qm_raw] if qm_raw else [])
        qm_display = ", ".join(qm_vals) if qm_vals else "—  (missing)"

        # what did v6 predict for this key?
        v6_raw = v6_attrs.get(key)
        v6_vals = v6_raw if isinstance(v6_raw, list) else ([v6_raw] if v6_raw else [])
        v6_display = ", ".join(v6_vals) if v6_vals else "—  (missing)"

        # match check: any overlap with gold values
        qm_match = bool(set(qm_vals) & set(gold_vals))
        v6_match = bool(set(v6_vals) & set(gold_vals))

        if qm_match: qm_correct += 1
        if v6_match: v6_correct += 1

        # write row
        ws.row_dimensions[row].height = 18
        cells_data = [
            (qi, center, GREY if qi % 2 == 0 else WHITE, normal),
            (query if row == first_row else "", left, QUERY_BG, bold_black),
            (key, left, WHITE, normal),
            (", ".join(gold_vals), left, WHITE, normal),
            (qm_display, left, WHITE, normal),
            (v6_display, left, WHITE, normal),
            ("✓ YES" if qm_match else "✗ NO", center, GREEN if qm_match else RED,
             bold_green if qm_match else bold_red),
            ("✓ YES" if v6_match else "✗ NO", center, GREEN if v6_match else RED,
             bold_green if v6_match else bold_red),
            ("", center, WHITE, normal),
        ]
        for ci, (val, align, fill, font) in enumerate(cells_data, 1):
            c = ws.cell(row, ci, val)
            style(c, fill=fill, font=font, align=align)

        row += 1

    # ── summary row for this query ──────────────────────────────────────────
    grand_qm_correct += qm_correct
    grand_v6_correct += v6_correct
    grand_total       += total_keys

    qm_pct = qm_correct / total_keys * 100
    v6_pct = v6_correct / total_keys * 100

    ws.row_dimensions[row].height = 20
    summary_cells = [
        ("", center, SUMMARY_BG, bold_black),
        (f"↳ QUERY TOTAL  ({total_keys} attributes)", left, SUMMARY_BG, bold_black),
        ("", center, SUMMARY_BG, bold_black),
        ("", center, SUMMARY_BG, bold_black),
        (f"qmeans: {qm_correct}/{total_keys}", center, SUMMARY_BG, bold_black),
        (f"V6: {v6_correct}/{total_keys}", center, SUMMARY_BG, bold_black),
        (f"{qm_pct:.0f}%", center, GREEN if qm_pct >= 60 else (YELLOW if qm_pct >= 30 else RED), bold_black),
        (f"{v6_pct:.0f}%", center, GREEN if v6_pct >= 60 else (YELLOW if v6_pct >= 30 else RED), bold_black),
        ("V6 better" if v6_correct > qm_correct else ("qmeans better" if qm_correct > v6_correct else "tie"),
         center, SUMMARY_BG, bold_black),
    ]
    for ci, (val, align, fill, font) in enumerate(summary_cells, 1):
        c = ws.cell(row, ci, val)
        style(c, fill=fill, font=font, align=align)

    row += 1

# ── grand total row ────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 25
gt_qm_pct = grand_qm_correct / grand_total * 100
gt_v6_pct = grand_v6_correct / grand_total * 100
grand_cells = [
    ("", center, BLUE_HDR, bold_white),
    (f"GRAND TOTAL  —  {len(gold)} queries, {grand_total} attributes", left, BLUE_HDR, bold_white),
    ("", center, BLUE_HDR, bold_white),
    ("", center, BLUE_HDR, bold_white),
    (f"qmeans: {grand_qm_correct}/{grand_total}", center, BLUE_HDR, bold_white),
    (f"V6: {grand_v6_correct}/{grand_total}", center, BLUE_HDR, bold_white),
    (f"{gt_qm_pct:.1f}%", center, BLUE_HDR, bold_white),
    (f"{gt_v6_pct:.1f}%", center, BLUE_HDR, bold_white),
    (f"V6 +{gt_v6_pct - gt_qm_pct:.1f}% better", center, BLUE_HDR, bold_white),
]
for ci, (val, align, fill, font) in enumerate(grand_cells, 1):
    c = ws.cell(row, ci, val)
    style(c, fill=fill, font=font, align=align)

# ── save ──────────────────────────────────────────────────────────────────────
out = 'bee/per_query_comparison.xlsx'
wb.save(out)
print(f"Saved: {out}")
print(f"Rows written: {row}")
print(f"qmeans overall: {grand_qm_correct}/{grand_total} = {gt_qm_pct:.1f}%")
print(f"V6     overall: {grand_v6_correct}/{grand_total} = {gt_v6_pct:.1f}%")
print(f"V6 is +{gt_v6_pct - gt_qm_pct:.1f}% better than qmeans")
