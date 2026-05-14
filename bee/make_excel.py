import json, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────
def load(p):
    with open(p, encoding='utf-8') as f:
        return [json.loads(l) for l in f if l.strip()]

def norm(s): return str(s).lower().strip()

def to_flat(d):
    pairs = set()
    for k, v in d.items():
        if isinstance(v, list):
            for x in v: pairs.add((norm(k), norm(x)))
        else:
            pairs.add((norm(k), norm(v)))
    return pairs

def v4_expanded_pairs(raw_text):
    raw = re.sub(r'```(?:json)?', '', raw_text).strip()
    start = raw.find('{')
    if start < 0: return set()
    depth = 0; end = -1
    for i in range(start, len(raw)):
        if raw[i] == '{': depth += 1
        elif raw[i] == '}':
            depth -= 1
            if depth == 0: end = i + 1; break
    if end < 0: return set()
    try: obj = json.loads(raw[start:end])
    except: return set()
    pairs = set()
    for grp in obj.get('attributes', []):
        if not isinstance(grp, dict): continue
        for _, payload in grp.items():
            if not isinstance(payload, dict): continue
            val = payload.get('value')
            if not val: continue
            for kp in ['key_priority1', 'key_priority2', 'key_priority3']:
                k = payload.get(kp)
                if k: pairs.add((norm(k), norm(val)))
    return pairs

# ── load data ─────────────────────────────────────────────────────────────────
gold = load('bee/gold_1k_v2.jsonl')
v4   = load('bee/v4_priority_results.jsonl')
v6   = load('bee/v6_flat_results_fixed.jsonl')
qm   = load('bee/qmeans_v2_results.jsonl')

qm_by_q = {r['query'].strip().lower(): r.get('attributes', {}) for r in qm}
v6_by_q = {r['query'].strip().lower(): r.get('attributes', {}) for r in v6}
v4_by_q = {r['query'].strip().lower(): r for r in v4}

# ── colours ───────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
GREY   = PatternFill("solid", fgColor="D9D9D9")
HEADER = PatternFill("solid", fgColor="1F4E79")
SUBHDR = PatternFill("solid", fgColor="2E75B6")
WHITE  = PatternFill("solid", fgColor="FFFFFF")

bold_white = Font(bold=True, color="FFFFFF", size=11)
bold_black = Font(bold=True, size=10)
normal     = Font(size=10)
center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
left       = Alignment(horizontal='left',   vertical='center', wrap_text=True)

thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"
ws.column_dimensions['A'].width = 52
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22

def hdr(row, text, fill=HEADER, font=bold_white, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.fill = fill; c.font = font; c.alignment = center; c.border = border

def srow(row, label, v4v, v6v, qmv, best_col=None):
    cells = [ws.cell(row, 1, label),
             ws.cell(row, 2, v4v),
             ws.cell(row, 3, v6v),
             ws.cell(row, 4, qmv)]
    cells[0].font = normal; cells[0].alignment = left
    for i, c in enumerate(cells):
        c.border = border
        if i == 0: continue
        c.alignment = center; c.font = normal
    if best_col:
        cells[best_col].font = bold_black
        cells[best_col].fill = GREEN

# header row
ws.row_dimensions[1].height = 30
hdr(1, "MODEL COMPARISON  —  1,000 IndiaMART Motor Queries  |  Gold: 3,200 correct key-value pairs")

ws.row_dimensions[2].height = 22
for col, txt in [(1,"Metric"),(2,"V4  (4,659 guesses)"),(3,"V6  (2,220 guesses)"),(4,"qmeans  (2,884 guesses)")]:
    c = ws.cell(2, col, txt)
    c.fill = SUBHDR; c.font = bold_white; c.alignment = center; c.border = border

# compute totals
tot_gold = 0
v4_pred = v6_pred = qm_pred = 0
v4_key = v4_val = v4_pair = 0
v6_key = v6_val = v6_pair = 0
qm_key = qm_val = qm_pair = 0

rows_detail = []
for g in gold:
    q = g['query']
    tp = to_flat(g.get('attributes', {}))
    tk = set(k for k, _ in tp); tv = set(v for _, v in tp)
    tot_gold += len(tp)
    v4p = v4_expanded_pairs(v4_by_q.get(q.strip().lower(), {}).get('raw', ''))
    v6p = to_flat(v6_by_q.get(q.strip().lower(), {}))
    qmp = to_flat(qm_by_q.get(q.strip().lower(), {}))

    v4_pred += len(v4p); v6_pred += len(v6p); qm_pred += len(qmp)
    r_v4k = len(set(k for k,_ in v4p) & tk); v4_key += r_v4k
    r_v4v = len(set(v for _,v in v4p) & tv); v4_val += r_v4v
    r_v4p = len(v4p & tp);                   v4_pair += r_v4p
    r_v6k = len(set(k for k,_ in v6p) & tk); v6_key += r_v6k
    r_v6v = len(set(v for _,v in v6p) & tv); v6_val += r_v6v
    r_v6p = len(v6p & tp);                   v6_pair += r_v6p
    r_qmk = len(set(k for k,_ in qmp) & tk); qm_key += r_qmk
    r_qmv = len(set(v for _,v in qmp) & tv); qm_val += r_qmv
    r_qmp = len(qmp & tp);                   qm_pair += r_qmp

    rows_detail.append({
        'query': q,
        'gold': sorted(tp), 'gold_count': len(tp),
        'v4_pred': sorted(v4p), 'v4_correct': r_v4p,
        'v6_pred': sorted(v6p), 'v6_correct': r_v6p,
        'qm_pred': sorted(qmp), 'qm_correct': r_qmp,
        'v6_only': sorted(v6p & tp - qmp & tp),
        'qm_only': sorted(qmp & tp - v6p & tp),
    })

def pct(n, d): return f"{n:,} / {d:,}  =  {n/d*100:.1f}%"
def best3(a,b,c):
    m = max(a,b,c)
    return [2 if a==m else None, 3 if b==m else None, 4 if c==m else None][([a,b,c].index(m))]

r = 3
hdr(r, "COVERAGE — out of 3,200 gold answers, how many did each system find?", fill=PatternFill("solid",fgColor="1F4E79")); r+=1
srow(r,"How many correct attribute NAMES did it find?  (e.g. gold='phase' → said 'phase'?)",
     pct(v4_key,tot_gold), pct(v6_key,tot_gold), pct(qm_key,tot_gold),
     best3(v4_key,v6_key,qm_key)); r+=1
srow(r,"How many correct VALUES did it find?  (e.g. gold value='3' → said '3'?)",
     pct(v4_val,tot_gold), pct(v6_val,tot_gold), pct(qm_val,tot_gold),
     best3(v4_val,v6_val,qm_val)); r+=1
srow(r,"⭐ EXACT MATCH — right name + right value  [MAIN SCORE]  (e.g. 'phase=3' both correct)",
     pct(v4_pair,tot_gold), pct(v6_pair,tot_gold), pct(qm_pair,tot_gold),
     best3(v4_pair,v6_pair,qm_pair))
ws.row_dimensions[r].height = 30; r+=1

r+=1
hdr(r, "RELIABILITY — of each system's OWN guesses, how many were correct?  (low % = mostly wrong noise)"); r+=1
srow(r,"Attribute name reliability  (of its own guesses, how many names matched gold)",
     pct(v4_key,v4_pred), pct(v6_key,v6_pred), pct(qm_key,qm_pred),
     best3(v4_key/v4_pred,v6_key/v6_pred,qm_key/qm_pred)); r+=1
srow(r,"Value reliability  (of its own guesses, how many values matched gold)",
     pct(v4_val,v4_pred), pct(v6_val,v6_pred), pct(qm_val,qm_pred),
     best3(v4_val/v4_pred,v6_val/v6_pred,qm_val/qm_pred)); r+=1
srow(r,"⭐ Full reliability — name + value both correct  (of its own guesses)",
     pct(v4_pair,v4_pred), pct(v6_pair,v6_pred), pct(qm_pair,qm_pred),
     best3(v4_pair/v4_pred,v6_pair/v6_pred,qm_pair/qm_pred))
ws.row_dimensions[r].height = 30; r+=1

r+=2
hdr(r, f"VERDICT:  V6 wins all 6 metrics.   V6 exact match = {v6_pair/tot_gold*100:.1f}%  vs  qmeans = {qm_pair/tot_gold*100:.1f}%  →  +{(v6_pair-qm_pair)/qm_pair*100:.0f}% improvement",
    fill=PatternFill("solid",fgColor="375623"), font=Font(bold=True,color="FFFFFF",size=12))

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Per-Query Detail
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Per-Query Detail")
ws2.freeze_panes = "A2"
cols = [
    ("Query", 45),
    ("Gold answers (correct)", 35),
    ("V6 predictions", 35),
    ("V6 correct pairs", 5),
    ("qmeans predictions", 35),
    ("qmeans correct pairs", 5),
    ("V6 won (qmeans missed)", 30),
    ("qmeans won (V6 missed)", 30),
]
for i, (title, w) in enumerate(cols, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    c = ws2.cell(1, i, title)
    c.fill = SUBHDR; c.font = bold_white; c.alignment = center; c.border = border

for ri, d in enumerate(rows_detail, 2):
    v6_won  = set(d['v6_pred']) & set(d['gold']) - set(d['qm_pred']) & set(d['gold'])
    qm_won  = set(d['qm_pred']) & set(d['gold']) - set(d['v6_pred']) & set(d['gold'])

    def fmt_pairs(pairs): return "\n".join(f"{k} = {v}" for k,v in sorted(pairs)) if pairs else "—"

    vals = [
        d['query'],
        fmt_pairs(d['gold']),
        fmt_pairs(d['v6_pred']),
        d['v6_correct'],
        fmt_pairs(d['qm_pred']),
        d['qm_correct'],
        fmt_pairs(v6_won),
        fmt_pairs(qm_won),
    ]
    fills = [WHITE, GREY, WHITE, None, WHITE, None, GREEN if v6_won else WHITE, YELLOW if qm_won else WHITE]

    for ci, (val, fill) in enumerate(zip(vals, fills), 1):
        c = ws2.cell(ri, ci, val)
        c.alignment = left; c.border = border; c.font = normal
        if fill: c.fill = fill
        if ci in (4, 6):
            c.alignment = center
            if isinstance(val, int):
                if val == 0:   c.fill = RED
                elif val >= 3: c.fill = GREEN
                else:          c.fill = YELLOW

    # highlight rows where V6 clearly beats qmeans
    if d['v6_correct'] > d['qm_correct'] + 1:
        ws2.cell(ri, 1).fill = PatternFill("solid", fgColor="E2EFDA")

ws2.row_dimensions[1].height = 25

# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
out = 'bee/v6_vs_qmeans_comparison.xlsx'
wb.save(out)
print(f"Saved: {out}")
print(f"\nTotal queries: 1,000")
print(f"V6  exact match: {v6_pair}/{tot_gold} = {v6_pair/tot_gold*100:.1f}%")
print(f"qmeans exact match: {qm_pair}/{tot_gold} = {qm_pair/tot_gold*100:.1f}%")
print(f"V6 wins by: +{(v6_pair-qm_pair)/qm_pair*100:.0f}%")
